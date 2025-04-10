import os
import re
import shutil
import math
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

# icon from: https://icons8.com
# author: Philipp Schaeffner


# *************************************************** GLOBALS *************************************************** #

local_dir = str(Path(__file__).resolve().parent)

import_dir = local_dir + '/IMPORT/'
export_dir = local_dir + '/EXPORT/'
export_excel_dir = export_dir + 'DONE.xlsx'
header_list = ['Comment', 'Date', 'Time', 'Duration', 'STIPA', 'LAeq', 'Status', 'LZeq', 'mr1', 'mr2', 'Status', 'LZeq', 'mr1', 'mr2', 'Status', 'LZeq', 'mr1', 'mr2', 'Status', 'LZeq', 'mr1', 'mr2', 'Status', 'LZeq', 'mr1', 'mr2', 'Status', 'LZeq', 'mr1', 'mr2', 'Status', 'LZeq', 'mr1', 'mr2', 'Status']


# ************************************************** FUNCTIONS ************************************************** #

def count_files(y):
    x = 0
    for path in os.listdir(y):
        if os.path.isfile(os.path.join(y, path)):
            x += 1


def printandquit_if_zero(y, z):
    if y == 0:
        print(z)
        input('Press any key to close...')
        exit()


def check_protokoll_type(y, z):
    if z not in y[0]:
        printandquit_if_zero(0,'ERROR: Falsches Protokoll Format!')

# ************************************************** MAIN LOOP ************************************************** #


number_of_imports = count_files(import_dir)

number_of_imports = 0
for path in os.listdir(import_dir):
    if os.path.isfile(os.path.join(import_dir, path)):
        number_of_imports += 1

printandquit_if_zero(number_of_imports, 'ERROR: IMPORT Pfad ist leer!')

number_of_sheets = math.ceil((number_of_imports - 3) / 3)


#x = 0-35; y = 0-n
data = [[0 for x in range(len(header_list))] for y in range(number_of_imports)]

# read data from all import files
count = 0
for filename in os.listdir(import_dir):                                          
    if filename.endswith('.txt'):                                               
        f = open(import_dir + filename, 'r')                             
        if f.mode == 'r':
            lines = f.readlines()

            check_protokoll_type(lines, 'STIPA')

            messurements = lines[18]
            messurements = re.sub(r'\s+', ';', messurements.strip())
            messurements_list = messurements.split(';')
            for i in range(len(header_list)):
                #print(str(i) + '_' + str(count) + '_' + str(len(messurements_list)))
                data[count][i] = messurements_list[i]
        f.close()
    count += 1



# create export xlsx and fitting number of worksheets
if os.path.exists(export_excel_dir):
    os.remove(export_excel_dir)
shutil.copyfile(local_dir + '/bin/' + 'BASE.xlsx', export_dir + 'DONE.xlsx')

workbook = load_workbook(filename=export_excel_dir)
mainsheet = workbook['STI_0']
for i in range(number_of_sheets):
    duplicate = workbook.copy_worksheet(mainsheet)
    duplicate.title = f'STI_{i+1}'
workbook.save(filename=export_excel_dir)



# fill sheets with data
workbook = load_workbook(filename=export_excel_dir)
y = 1
for i in range(number_of_sheets + 1):
    activesheet = workbook['STI_' + str(i)]

    if y <= number_of_imports:
        activesheet.cell(row=2, column=4).value  = data[y-1][0]      #Posi_Name 1
        activesheet.cell(row=2, column=3).value  = y                 #Posi_Nummer
        activesheet.cell(row=4, column=3).value  = ''                #Posi_Lage
        activesheet.cell(row=8, column=3).value  = data[y-1][7]      #125Hz
        activesheet.cell(row=8, column=4).value  = data[y-1][11]     #250Hz
        activesheet.cell(row=8, column=5).value  = data[y-1][15]     #500Hz
        activesheet.cell(row=8, column=6).value  = data[y-1][19]     #1kHz
        activesheet.cell(row=8, column=7).value  = data[y-1][23]     #2kHz
        activesheet.cell(row=8, column=8).value  = data[y-1][27]     #4kHz
        activesheet.cell(row=8, column=9).value  = data[y-1][31]     #8kHz
        activesheet.cell(row=8, column=10).value  = data[y-1][5]     #Summenpegel
        activesheet.cell(row=13, column=3).value  = data[y-1][8]     #Verst. 0
        activesheet.cell(row=13, column=4).value  = data[y-1][12]    #Verst. 1
        activesheet.cell(row=13, column=5).value  = data[y-1][16]    #Verst. 2
        activesheet.cell(row=13, column=6).value  = data[y-1][20]    #Verst. 3
        activesheet.cell(row=13, column=7).value  = data[y-1][24]    #Verst. 4
        activesheet.cell(row=13, column=8).value  = data[y-1][28]    #Verst. 5
        activesheet.cell(row=13, column=9).value  = data[y-1][32]    #Verst. 6
        activesheet.cell(row=13, column=10).value  = data[y-1][4]    #Mittelwert
        activesheet.cell(row=14, column=3).value  = data[y-1][6]     #Anmerkung/Status
        y += 1 

    if y <= number_of_imports:
        activesheet.cell(row=16, column=4).value = data[y-1][0]      #Posi_Name 1
        activesheet.cell(row=16, column=3).value = y                 #Posi_Nummer
        activesheet.cell(row=18, column=3).value  = ''               #Posi_Lage
        activesheet.cell(row=22, column=3).value  = data[y-1][7]     #125Hz
        activesheet.cell(row=22, column=4).value  = data[y-1][11]    #250Hz
        activesheet.cell(row=22, column=5).value  = data[y-1][15]    #500Hz
        activesheet.cell(row=22, column=6).value  = data[y-1][19]    #1kHz
        activesheet.cell(row=22, column=7).value  = data[y-1][23]    #2kHz
        activesheet.cell(row=22, column=8).value  = data[y-1][27]    #4kHz
        activesheet.cell(row=22, column=9).value  = data[y-1][31]    #8kHz
        activesheet.cell(row=22, column=10).value  = data[y-1][5]    #Summenpegel
        activesheet.cell(row=27, column=3).value  = data[y-1][8]     #Verst. 0
        activesheet.cell(row=27, column=4).value  = data[y-1][12]    #Verst. 1
        activesheet.cell(row=27, column=5).value  = data[y-1][16]    #Verst. 2
        activesheet.cell(row=27, column=6).value  = data[y-1][20]    #Verst. 3
        activesheet.cell(row=27, column=7).value  = data[y-1][24]    #Verst. 4
        activesheet.cell(row=27, column=8).value  = data[y-1][28]    #Verst. 5
        activesheet.cell(row=27, column=9).value  = data[y-1][32]    #Verst. 6
        activesheet.cell(row=27, column=10).value  = data[y-1][4]    #Mittelwert
        activesheet.cell(row=28, column=3).value  = data[y-1][6]     #Anmerkung/Status
        y += 1 

    if y <= number_of_imports:    
        activesheet.cell(row=30, column=4).value = data[y-1][0]      #Posi_Name 1
        activesheet.cell(row=30, column=3).value = y                 #Posi_Nummer
        activesheet.cell(row=32, column=3).value  = ''               #Posi_Lage
        activesheet.cell(row=36, column=3).value  = data[y-1][7]     #125Hz
        activesheet.cell(row=36, column=4).value  = data[y-1][11]    #250Hz
        activesheet.cell(row=36, column=5).value  = data[y-1][15]    #500Hz
        activesheet.cell(row=36, column=6).value  = data[y-1][19]    #1kHz
        activesheet.cell(row=36, column=7).value  = data[y-1][23]    #2kHz
        activesheet.cell(row=36, column=8).value  = data[y-1][27]    #4kHz
        activesheet.cell(row=36, column=9).value  = data[y-1][31]    #8kHz
        activesheet.cell(row=36, column=10).value  = data[y-1][5]    #Summenpegel
        activesheet.cell(row=41, column=3).value  = data[y-1][8]     #Verst. 0
        activesheet.cell(row=41, column=4).value  = data[y-1][12]    #Verst. 1
        activesheet.cell(row=41, column=5).value  = data[y-1][16]    #Verst. 2
        activesheet.cell(row=41, column=6).value  = data[y-1][20]    #Verst. 3
        activesheet.cell(row=41, column=7).value  = data[y-1][24]    #Verst. 4
        activesheet.cell(row=41, column=8).value  = data[y-1][28]    #Verst. 5
        activesheet.cell(row=41, column=9).value  = data[y-1][32]    #Verst. 6
        activesheet.cell(row=41, column=10).value  = data[y-1][4]    #Mittelwert
        activesheet.cell(row=42, column=3).value  = data[y-1][6]     #Anmerkung/Status
        y += 1 

workbook.save(filename=export_excel_dir)


printandquit_if_zero(0, 'DONE.xlsx wurde Erfolgreich erstellt!')
    
